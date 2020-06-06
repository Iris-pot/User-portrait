import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pyecharts import options as opts
from pyecharts.charts import Radar, Bar, Line, make_snapshot, Map, Page, Funnel
from pyecharts.faker import Faker
from wordcloud import wordcloud
import jieba
import imageio
import pygal
import openpyxl as xl
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

#女性居住地
prov = pd.read_csv("data.csv", encoding='gb18030')
da_1 = [[i, j] for i, j in zip(prov['area'], prov['cnt1'])]
map_graph = (Map().add("cnt1", da_1, "china").set_global_opts(
    title_opts=opts.TitleOpts(title="女性居住地分布情况"),
    visualmap_opts=opts.VisualMapOpts(max_=100)))
map_graph.render('area.html')

#择偶要求首选居住地
da_2 = [[i, j] for i, j in zip(prov['sarea'], prov['cnt2'])]
map_graph = (Map().add("cnt2", da_1, "china").set_global_opts(
    title_opts=opts.TitleOpts(title="女性居住地分布情况"),
    visualmap_opts=opts.VisualMapOpts(max_=100)))
map_graph.render('sarea.html')

#相亲女性自我介绍
mk = imageio.imread('model.jpeg')
w = wordcloud.WordCloud(width=1000,
                        height=800,
                        background_color='white',
                        font_path='msyh.ttc',
                        mask=mk,
                        scale=20)
f = open(r'C:\Users\YFLY\可视化数据\introduction.txt', encoding='utf8')
txt = f.read()
txtlist = jieba.lcut(txt)
string = ' '.join(txtlist)
w.generate(string)
w.to_file(r'C:\Users\YFLY\可视化数据\introduction_wordcloud.jpg')

#女性薪资分布情况
df = pd.read_csv('data.csv', encoding='GBK')
x = list(df['salary'])[:6]
y = list(df['cnt3'])[:6]
for i in range(5):
    plt.scatter(x[i], y[i], s=y[i] * 10)
plt.grid(axis='y')
plt.savefig('aa.jpg', dpi=300)
plt.show()

#女性专业职业分布情况
data = pd.read_csv("data.csv", encoding='gb18030')
major = [['管理类', 105], ['艺术类', 78], ['经济学类', 72], ['医学类', 49], ['中文类', 44],
         ['教育类', 39], ['金融学类', 39], ['外文类', 38], ['市场营销类', 24], ['法学类', 21],
         ['计算机类', 13], ['电子信息类', 12], ['机械类', 10], ['建筑类', 10]]
Funnel = (Funnel().add("data",
                       major,
                       label_opts=opts.LabelOpts(position='inside')))
Funnel.render("major.html")
data = pd.read_csv("data.csv", encoding='gb18030')
pro = [['教育/培训', 106], ['人力资源/行政/后勤', 67], ['财会/审计/统计', 55], ['高级管理', 52],
       ['医疗/护理', 42], ['公务员/国家干部', 40], ['自由职业者', 40], ['私营业主', 33],
       ['金融/证券/投资/保险', 27], ['文学/传媒/影视', 27], ['销售', 25], ['采购/贸易', 20],
       ['设计/创意', 20], ['房地产/装修/物业', 18]]
Funnel = (Funnel().add("data",
                       major,
                       label_opts=opts.LabelOpts(position='inside')))
Funnel.render("pro.html")

#女性年龄分布
circle = pygal.Pie(inner_radius=0.6)
circle.title = '年龄分布图'
circle.add("21-30", 6.49)
circle.add('31-40', 36.67)
circle.add('41-50', 32.67)
circle.add('51-60', 18.40)
circle.add('61-', 5.78)
circle.render_to_file('age.svg')

#择偶年龄要求
test = xl.load_workbook(r'age.xlsx')
test_1 = test['sheet1']
for i in range(2, test_1.max_row + 1):
    s = []
    for j in range(2, test_1.max_row + 1):
        if test_1.cell(row=j, column=1).value == test_1.cell(row=i,
                                                             column=3).value:
            s.append(test_1.cell(row=j, column=2).value)
        if s != []:
            ss = sum(s) / len(s)
        else:
            ss = 0
    test_1.cell(row=i, column=4).value = ss
test.save('age.xlsx')
age = pd.read_csv("age.csv", encoding='gb18030')
plt.figure(figsize=(9, 3.5))
x = age['s']
y = age['期望男性最大年龄']
z = age['期望男性最小年龄']
w = age['女性年龄']
plt.plot(x, w, label="女性年龄")
plt.plot(x, y, label='期望男性最大年龄')
plt.plot(x, z, label='期望男性最小年龄')
plt.legend(loc='best')
plt.ylim([0, 100])
plt.xticks(np.linspace(24, 76, 27))

#身高分布
data = pd.read_csv("data.csv", encoding='gb18030')
plt.figure(figsize=(6, 3.5))
h = list(data['height'])[:6]
c = list(data['cnt8'])[:6]
plt.fill_between(h, c, color='blue', alpha=0.3)

test = xl.load_workbook(r'height.xlsx')
test_1 = test['height']
for i in range(2, test_1.max_row + 1):
    s = []
    for j in range(2, test_1.max_row + 1):
        if test_1.cell(row=j, column=1).value == test_1.cell(row=i,
                                                             column=3).value:
            s.append(test_1.cell(row=j, column=2).value)
        if s != []:
            ss = sum(s) / len(s)
        else:
            ss = 0
    test_1.cell(row=i, column=4).value = ss

test.save('height.xlsx')

height = pd.read_csv("height.csv", encoding='gb18030')
plt.figure(figsize=(9, 3.5))
x = height['sg']
y = height['期望男性最低身高']
z = height['女性身高']
plt.plot(x, z, label="女性身高")
plt.plot(x, y, label='期望男性最低身高')
plt.legend(loc='best')
plt.ylim([135, 185])
plt.xticks(np.linspace(150, 176, 27))

#学历分布
data = pd.read_csv("data.csv", encoding='gb18030')
labels = list(data['education'])[:5]
plt.figure(figsize=(9, 9))
sizes = list(data['cnt9'])[:5]
colors = ['red', 'blue', 'purple', 'green', 'gold']
plt.pie(sizes, labels=labels, autopct='%1.1f%%')
plt.axis('equal')
plt.savefig('./学历')
plt.show()

#女性婚姻状况
data = pd.read_csv("data.csv", encoding='gb18030')
plt.figure(figsize=(5, 5))
marry = list(data['marry'])[:3]
c = list(data['cnt10'])[:3]
plt.pie(c, labels=marry, autopct='%1.1f%%')
plt.savefig('./marry')

#理想婚姻
mk = imageio.imread('ideal.jpeg')
w = wordcloud.WordCloud(width=1000,
                        height=800,
                        background_color='white',
                        font_path='msyh.ttc',
                        mask=mk,
                        scale=20)
f = open('C:\Users\YFLY\可视化数据\ideal-marriage.txt', encoding='utf-8')
txt = f.read()
txtlist = jieba.lcut(txt)
string = ' '.join(txtlist)
w.generate(string)
w.to_file('C:\Users\YFLY\可视化数据\ideal-marriage_wordcloud.jpg')

#是否愿意要小孩
data = pd.read_csv("data.csv", encoding='gb18030')
plt.figure(figsize=(7, 5))
child = list(data['child'])[:4]
c = list(data['cnt11'])[:4]
plt.bar(child, c, color='y', alpha=0.6)
plt.savefig('./child')

#择偶婚姻要求
from pyecharts import options as opts
from pyecharts.charts import Radar
value_wh = [[152, 1, 2, 175, 62, 6, 0]]
value_so = [[20, 0, 13, 0, 0, 1, 2]]
value_ly = [[248, 31, 99, 2, 0, 33, 1]]
c_schema = [
    {
        "name": '不限',
        'max': 250,
        'min': 0
    },
    {
        "name": '离异',
        'max': 250,
        'min': 0
    },
    {
        "name": '离异、丧偶',
        'max': 250,
        'min': 0
    },
    {
        "name": '未婚',
        'max': 250,
        'min': 0
    },
    {
        "name": '未婚、离异',
        'max': 250,
        'min': 0
    },
    {
        "name": '未婚、丧偶',
        'max': 250,
        'min': 0
    },
    {
        "name": '丧偶',
        'max': 250,
        'min': 0
    },
]
Radar = (Radar().add_schema(schema=c_schema, shape='circle').add(
    '未婚', value_wh, color='#f9713c').add('丧偶', value_so, color='#b3e4a1').add(
        '离异', value_ly, color='#00da3c').set_series_opts(
            label_opts=opts.LabelOpts(is_show=False)))
Radar.render('择偶婚姻需求.html')

#不同年龄段的理想结婚时间
bar = (
    Bar().add_xaxis([
        "21岁-30岁",
        "31岁-40岁",
        "41岁-50岁",
        "51岁-60岁",
        "61岁以上",
    ]).add_yaxis("尽快结婚", [4.88, 19.71, 20.11, 9.91, 6.45]).add_yaxis(
        "时机成熟就结婚",
        [60.98, 49.04, 60.89, 73.87, 83.87]).add_yaxis("一年内结婚", [
            29.27, 30.29, 16.76, 15.32, 6.45
        ]).add_yaxis("两年内结婚", [4.88, 0.96, 1.68, 0.90, 3.23]).add_yaxis(
            "三年内不考虑结婚",
            [0, 0, 0.56, 0, 0]).set_global_opts(
                yaxis_opts=opts.AxisOpts(
                    name="百分百",
                    type_="value",
                    axislabel_opts=opts.LabelOpts(formatter="{value} %"),
                    axistick_opts=opts.AxisTickOpts(is_show=True),
                    splitline_opts=opts.SplitLineOpts(is_show=True),
                )
                #        title_opts=opts.TitleOpts(title="Bar-旋转X轴标签", subtitle="解决标签名字过长的问题"),
            ))
bar.render("不同年龄段的理想结婚时间.html")

#不同学历对小孩的需求占比
bar1 = (Bar().add_xaxis([
    "博士",
    "硕士",
    "本科",
    "大专",
    "高中中专及以下",
]).add_yaxis(
    "不愿意要孩子", [11.11, 5.45, 22.25, 40.98, 44.05], stack="stack1").add_yaxis(
        "要不要孩子都行", [11.11, 3.64, 1.31, 3.38, 0], stack="stack1").add_yaxis(
            "要不要孩子视情况而定", [11.11, 14.55, 21.20, 24.44, 26.19],
            stack="stack1").add_yaxis(
                "愿意要孩子", [66.67, 76.36, 55.24, 31.20, 29.76],
                stack="stack1").reversal_axis().set_global_opts(
                    xaxis_opts=opts.AxisOpts(
                        name="百分百",
                        type_="value",
                        axislabel_opts=opts.LabelOpts(formatter="{value} %"),
                        axistick_opts=opts.AxisTickOpts(is_show=True),
                        splitline_opts=opts.SplitLineOpts(is_show=True),
                    )).set_series_opts(label_opts=opts.LabelOpts(
                        is_show=False)))